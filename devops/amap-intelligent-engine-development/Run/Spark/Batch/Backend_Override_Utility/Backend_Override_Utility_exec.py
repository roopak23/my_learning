import traceback
import sys
import os
import boto3
import pandas as pd
from io import BytesIO
import pymysql
import openpyxl
from sqlalchemy import create_engine, text
from sys import argv


def execute_sql_script(sql_query):
    global connection_properties
    # Save the SQL query to a temporary file
    current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    random_number = random.randint(1000, 9999)
    temp_sql_file = f"temp_sql_script_{current_time}_{random_number}.sql"
    print(f"Generated SQL file name: {temp_sql_file}")
    try:
        # Write the SQL query to the temporary file
        with open(temp_sql_file, 'w') as file:
            file.write(sql_query)
        print(f"SQL script saved to {temp_sql_file}.")
        # Prepare the MySQL command
        mysql_command = [
            'mysql',
            '-u', connection_properties['user'],
            '-p' + connection_properties['password'],
            '-h', connection_properties['host'],
            connection_properties['database'],
            '-e', f"source {temp_sql_file}"
        ]
        # Execute the SQL script using MySQL client
        subprocess.run(mysql_command, check=True)
        print("SQL script executed successfully using MySQL client.")
        return True 
    except subprocess.CalledProcessError as e:
        print(f"Error executing SQL script via MySQL client: {e}")
        return False 
    except Exception as e:
        print(f"Error handling temporary SQL file: {e}")
        return False
    finally:
        # Clean up the temporary SQL file
        if os.path.exists(temp_sql_file):
            #os.remove(temp_sql_file)
            print(f"Temporary SQL file {temp_sql_file} is not removing.")

def update_excel_column_in_s3(bucket_name, file_key, sheet_name, column_name, new_values=None, updated_file_key=None):
    #print("printing new values :::", False if not new_values else True)
    # Initialize S3 client
    s3 = boto3.client('s3')
    # Fetch the Excel file from S3
    response = s3.get_object(Bucket=bucket_name, Key=file_key)
    file_content = response['Body'].read()
    # Load the Excel file into a DataFrame, specifying the sheet
    with BytesIO(file_content) as excel_data:
        df = pd.read_excel(excel_data, sheet_name=sheet_name, engine='openpyxl')
    # Check if the column exists in the specified sheet's DataFrame
    if column_name not in df.columns:
        raise ValueError(f"Column '{column_name}' not found in the sheet '{sheet_name}' of the Excel file.")
    # Update the specified column with new values
    if not new_values:
        #raise ValueError("The length of new_values does not match the number of rows in the Excel sheet.")
        print("Reading the control files...")
        print("Returning the contol file data ::: , ")
        print(df)
        print('\n')
    else:
        df[column_name] = new_values
        # Save the updated DataFrame back to Excel in memory
        with BytesIO() as output:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Load all sheets, update the specified sheet, and save back to Excel
                with pd.ExcelFile(BytesIO(file_content), engine='openpyxl') as xls:
                    for sheet in xls.sheet_names:
                        if sheet == sheet_name:
                            df.to_excel(writer, sheet_name=sheet, index=False)
                        else:
                            pd.read_excel(xls, sheet_name=sheet).to_excel(writer, sheet_name=sheet, index=False)
            output.seek(0)
            # Upload the updated file back to S3 with appropriate ContentType
            updated_file_key = updated_file_key or file_key
            s3.put_object(
                Bucket=bucket_name,
                Key=updated_file_key,
                Body=output.getvalue(),
                ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        print(f"Successfully updated '{column_name}' in sheet '{sheet_name}' of '{updated_file_key}' on S3 bucket '{bucket_name}'.")
    return df.iloc[0]['Tag'], df.iloc[0]['Processing_Status'], df.iloc[0]['Start_Date'], df.iloc[0]['End_Date'], df.iloc[0]['Level1'], df.iloc[0]['Level2'], df.iloc[0]['Level3'], df.iloc[0]['Level4'], df.iloc[0]['Level5'], df.iloc[0]['Event'], df.iloc[0]['Overwriting_reason'], df.iloc[0]['Use_Overwrite'], df.iloc[0]['overwritten_impression_number'] ,df 

def update_excel_in_s3(bucket_name, file_key, df, sheet_name):
    s3_client = boto3.client('s3')
    try:
        # Get the existing Excel file from S3
        s3_object = s3_client.get_object(Bucket=bucket_name, Key=file_key)
        file_stream = s3_object['Body'].read()
        # Load the Excel file into an OpenPyXL workbook
        workbook = openpyxl.load_workbook(BytesIO(file_stream))
        # Check if the specified sheet exists, if not, raise an error
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' does not exist in the Excel file.")
            return
        # Get the specified sheet
        sheet = workbook[sheet_name]
        # Clear existing content in the sheet
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            for cell in row:
                cell.value = None
        # Write the updated DataFrame to the sheet
        for i, column in enumerate(df.columns, start=1):
            sheet.cell(row=1, column=i, value=column)  # Set header row
        for i, row in enumerate(df.values, start=2):
            for j, value in enumerate(row, start=1):
                sheet.cell(row=i, column=j, value=value)
        # Save the updated workbook to a BytesIO object
        updated_file_stream = BytesIO()
        workbook.save(updated_file_stream)
        updated_file_stream.seek(0)
        # Upload the updated file back to S3
        s3_client.put_object(Bucket=bucket_name, Key=file_key, Body=updated_file_stream)
        print(f"Successfully updated the '{sheet_name}' sheet in the Excel file.")
    except Exception as e:
        print(f"An error occurred while updating the Excel file: {e}")


def list_sheet_names(bucket_name, file_key):
    s3_client = boto3.client('s3')
    response = s3_client.get_object(Bucket=bucket_name, Key=file_key)
    excel_data = response['Body'].read()
    with BytesIO(excel_data) as excel_file:
        xls = pd.ExcelFile(excel_file)
        return xls.sheet_names

def generate_download_url(bucket_name, file_key, expiration=3600):
    s3_client = boto3.client('s3')
    try:
        # Generate a pre-signed URL to allow downloading the file from S3
        url = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': bucket_name, 'Key': file_key},
            ExpiresIn=expiration
        )
        print(f"Updated Excel can be pulled using the following URL: {url}")
        return url
    except Exception as e:
        print(f"An error occurred while generating the download URL: {e}")
        return None

def create_rds_connection(connection_properties):
    try:
        # Establish the connection using pymysql
        connection = pymysql.connect(
            host=connection_properties["host"],
            user=connection_properties["user"],
            password=connection_properties["password"],
            database=connection_properties["database"],
            port=connection_properties["port"]
        )
        print("Connection to RDS database successful")
        return connection
    except pymysql.MySQLError as e:
        print(f"Error connecting to RDS database: {e}")
        return None

def update_levels_event_combinations(connection):
    try:
        global tag
        sheet_name = 'input'
        column_name = 'Processing_Status'
        new_values = ['running']  # values to update in the column
        with connection.cursor() as cur:
            # Sample query - replace with your actual query
            if tag[0] == 'Active_Levels_and_Event' and tag[1] == 'in-progress':
                # updating latest status of input sheet
                tag = update_excel_column_in_s3(bucket_name, file_key, sheet_name,  column_name, new_values, updated_file_key)
                query = "select distinct Level1,Level2,Level3,Level4,Level5,ai.event as Event from (select distinct adserver_adslot_id,event from api_inventorycheck) ai left join STG_SA_Ad_Slot ssas on ai.adserver_adslot_id =ssas.adserver_adslot_id where ssas.status ='ACTIVE'"
                df = pd.read_sql(query, connection)
                print(f"Record count: {df.shape[0]}")
                sheet_name = 'Active_Levels_and_Event'
                update_excel_in_s3(bucket_name, file_key, df, sheet_name)
                print(f"Updated Excel can be pulled using the following URL: https://{bucket_name}.s3.ap-southeast-2.amazonaws.com/{file_key}")
                # updating latest status of input sheet
                sheet_name = 'input'
                column_name = 'Processing_Status'
                new_values = ['completed']  # values to update in the column
                tag = update_excel_column_in_s3(bucket_name, file_key, sheet_name,  column_name, new_values, updated_file_key)
            else:
                print("No request to provide updated active Levels and Event Combinations.")
    except pymysql.MySQLError as e:
        print(f"Error fetching record count: {e}")
        return None

def read_near_realtime_levels_event_combinations_capacity(connection):
    try:
        global tag
        with connection.cursor() as cur:
            # Sample query - replace with your actual query
            if tag[0] == 'Report' and tag[1] == 'in-progress':
                # updating latest status of input sheet
                sheet_name = 'input'
                column_name = 'Processing_Status'
                new_values = ['running']  # values to update in the column
                tag = update_excel_column_in_s3(bucket_name, file_key, sheet_name,  column_name, new_values, updated_file_key)
                query = "select distinct Level1,Level2,Level3,Level4,Level5,ai.event as Event from (select distinct adserver_adslot_id,event from api_inventorycheck) ai left join STG_SA_Ad_Slot ssas on ai.adserver_adslot_id =ssas.adserver_adslot_id where ssas.status ='ACTIVE'"
                df = pd.read_sql(query, connection)
                print(f"Record count: {df.shape[0]}")
                sheet_name = 'Active_Levels_and_Event'
                update_excel_in_s3(bucket_name, file_key, df, sheet_name)
                print(f"Updated Excel can be pulled using the following URL: https://{bucket_name}.s3.ap-southeast-2.amazonaws.com/{file_key}")
                # updating latest status of input sheet
                sheet_name = 'input'
                column_name = 'Processing_Status'
                new_values = ['completed']  # values to update in the column
                tag = update_excel_column_in_s3(bucket_name, file_key, sheet_name,  column_name, new_values, updated_file_key)
            else:
                print(f"No request to update Excel Sheet, {sheet_name}")
    except pymysql.MySQLError as e:
        print(f"Error fetching record count: {e}")
        return None

def construct_query(start_date, end_date, level1=None, event=None, levels=None):
    # Check for required parameters
    if not start_date or not end_date or (not level1 and not event):
        #return "Required fields missing: start_date, end_date, and either level1 or event must be provided."
        raise ValueError("Required fields missing: start_date, end_date, and either level1 or event must be provided.")
    # Ensure mandatory columns are included
    select_columns = ["`date`","use_overwrite", "overwritten_expiry_date", "Level1"]
    group_by_columns = ["`date`","use_overwrite", "overwritten_expiry_date", "Level1"]
    # Dynamically add levels if provided
    if levels:
        for i, level in enumerate(levels, start=2):
            col_name = f"Level{i}"
            if pd.notna(level):  # Check if level is not NaN
                select_columns.append(col_name)
                group_by_columns.append(col_name)
            else:
                select_columns.append(f"'' AS {col_name}")
    # Handle the event column separately
    if event and pd.notna(event):
        select_columns.append("event AS Event")
        group_by_columns.append("event")
    else:
        select_columns.append("'' AS event")
    # Construct the select and group by clauses dynamically
    select_clause = ", ".join(select_columns)
    group_by_clause = ", ".join(group_by_columns)
    # Start forming the base query
    query = f"""
    SELECT 
        {select_clause},
        SUM(future_capacity) AS future_capacity,
        SUM(overwritten_impressions) AS overwritten_impressions, 
        SUM(CASE WHEN use_overwrite = 'Y' THEN overwritten_impressions ELSE future_capacity END) AS capacity_considered_in_pengana,
        SUM(booked) AS booked,
        SUM(reserved) AS reserved
    FROM (
        SELECT 
            ai.`date`,
            ai.use_overwrite,
            ai.overwritten_expiry_date,
            ai.future_capacity,
            ai.overwritten_impressions,
            ai.booked,
            ai.reserved,
            ssas.Level1,
            ssas.Level2,
            ssas.Level3,
            ssas.Level4,
            ssas.Level5,
            ai.event
        FROM 
            api_inventorycheck ai 
            INNER JOIN STG_SA_Ad_Slot ssas ON ai.adserver_adslot_id = ssas.adserver_adslot_id
        WHERE 
            ai.date BETWEEN '{start_date}' AND '{end_date}'
    """
    # Add conditions based on dynamic parameters
    if level1:
        query += f" AND ssas.Level1 = '{level1}'"
    if levels:
        for i, level in enumerate(levels, start=2):
            if pd.notna(level):  # Only add the condition if the level is not NaN
                query += f" AND ssas.Level{i} = '{level}'"
    if event and pd.notna(event):
        query += f" AND ai.event = '{event}'"
    # Add status filter
    query += " AND ssas.status = 'ACTIVE'"
    # Close subquery and add group by clause
    query += f"""
    ) a
    GROUP BY {group_by_clause}
    """
    return query

def construct_query_pre(connection,bucket_name, file_key, sheet_name, column_name, new_values, updated_file_key):
    global tag
    if tag[0] == 'Report' and tag[1] == 'in-progress':
        sheet_name = 'input'
        column_name = 'Processing_Status'
        new_values = ['running']  # values to update in the column
        tag = update_excel_column_in_s3(bucket_name, file_key, sheet_name, column_name, new_values, updated_file_key)
        # Sample usage of the query function
        query = construct_query(
            start_date=tag[2],
            end_date=tag[3],
            level1=tag[4],
            event=tag[9],
            levels=[tag[i] for i in range(5, 9)]
        )
        if "Required fields missing" in query:
            print(query)
        else:
            query_df = pd.read_sql(query, connection)
            print(f"Record count for Read tag: {query_df.shape[0]}")
            # Print the generated query for demonstration purposes
            print("Generated Query:", query)
            sheet_name = 'report'
            update_excel_in_s3(bucket_name, file_key, query_df, sheet_name)
            print(f"Updated Excel can be pulled using the following URL: https://{bucket_name}.s3.ap-southeast-2.amazonaws.com/{file_key}")
            sheet_name = 'input'
            column_name = 'Processing_Status'
            new_values = ['Completed']  # values to update in the column
            tag = update_excel_column_in_s3(bucket_name, file_key, sheet_name,  column_name, new_values, updated_file_key)

def clean_param(value):
    if pd.isna(value) or value == "":
        return None
    return value

def update_overwrite_for_combinations(connection,bucket_name, file_key, sheet_name, column_name, new_values, updated_file_key):
    try:
        global tag
        # Sample query - replace with your actual query
        df = tag[13]
        if not tag[2] or not tag[3] or (not tag[4] and not tag[9]):
            #return "Required fields missing: start_date, end_date, and either level1 or event must be provided."
            raise ValueError("Required fields missing: start_date, end_date, and either level1 or event must be provided.")
        else:
            if tag[0] == 'Update' and tag[1] == 'in-progress' and df.iloc[0]['overwritten_impression_number'] >= 0:
                # updating latest status of input sheet
                #sql_script = 's3://foxtel-hawkeye-uat-s3-data/data/inbound/override_backend_utility_code/backend_utility_sql_code_execute.sql'
                #sql_script_download_command = f"aws s3 cp {sql_script} sql_script.sql"
                #os.system(sql_script_download_command)
                os.system("cp -f /tmp/Backend_Override_Utility/backend_utility_sql_code_execute.sql /tmp/Backend_Override_Utility/sql_script.sql")
                sheet_name = 'input'
                column_name = 'Processing_Status'
                new_values = ['running']  # values to update in the column
                tag = update_excel_column_in_s3(bucket_name, file_key, sheet_name,  column_name, new_values, updated_file_key)
                params = {
                    'overwritten_impression_number': int(clean_param(df.iloc[0]['overwritten_impression_number'])) if pd.notna(df.iloc[0]['overwritten_impression_number']) else None,
                    'level1': str(clean_param(df.iloc[0]['Level1'])) if pd.notna(df.iloc[0]['Level1']) else None,
                    'level2': str(clean_param(df.iloc[0]['Level2'])) if pd.notna(df.iloc[0]['Level2']) else None,
                    'level3': str(clean_param(df.iloc[0]['Level3'])) if pd.notna(df.iloc[0]['Level3']) else None,
                    'level4': str(clean_param(df.iloc[0]['Level4'])) if pd.notna(df.iloc[0]['Level4']) else None,
                    'level5': str(clean_param(df.iloc[0]['Level5'])) if pd.notna(df.iloc[0]['Level5']) else None,
                    'event': str(clean_param(df.iloc[0]['Event'])) if pd.notna(df.iloc[0]['Event']) else None,
                    'startDate': str(pd.to_datetime(df.iloc[0]['Start_Date']).strftime('%Y-%m-%d')) if pd.notna(df.iloc[0]['Start_Date']) else None,
                    'endDate': str(pd.to_datetime(df.iloc[0]['End_Date']).strftime('%Y-%m-%d')) if pd.notna(df.iloc[0]['End_Date']) else None,
                    'Use_Overwrite': str(clean_param(df.iloc[0]['Use_Overwrite'])) if pd.notna(df.iloc[0]['Use_Overwrite']) else None,
                    'Overwriting_reason': str(clean_param(df.iloc[0]['Overwriting_reason'])) if pd.notna(df.iloc[0]['Overwriting_reason']) else None
                }
                print("below parameters are passed to SQL script : {}".format(params))
                with open('/tmp/Backend_Override_Utility/sql_script.sql', 'r') as file:
                    sql_query = file.read()
                # Replace placeholders in the SQL script with actual parameter values
                for key, value in params.items():
                    # Replacing placeholders like #overwritten_impression_number# with actual values
                    sql_query = sql_query.replace(f"#{key}#",str(value) if key == 'overwritten_impression_number' and value is not None else f"'{value}'" if value is not None else 'NULL')
                # Execute the updated SQL query
                #print("SQL script : {}".format(sql_query))
                print("::: Start running the script :::")
                script_status = execute_sql_script(sql_query)  # Capture the status of the execution
                #print("::: Cheers...! Script is Successfully completed :::" if script_status else "::: Failed to complete script :::")
                if script_status:
                    print("::: Cheers...! Script is Successfully completed :::")
                    sheet_name = 'input'
                    column_name = 'Processing_Status'
                    new_values = ['in-progress']  # values to update in the column
                    tag = update_excel_column_in_s3(bucket_name, file_key, sheet_name,  column_name, new_values, updated_file_key)
                    print(tag)
                    column_name = 'Tag'
                    new_values = ['Report']  # values to update in the column
                    tag = update_excel_column_in_s3(bucket_name, file_key, sheet_name,  column_name, new_values, updated_file_key)
                    print(tag)
                    construct_query_pre(connection,bucket_name, file_key, sheet_name, column_name, new_values, updated_file_key)
                    column_name = 'Processing_Status'
                    new_values = ['Completed']  # values to update in the column
                    tag = update_excel_column_in_s3(bucket_name, file_key, sheet_name,  column_name, new_values, updated_file_key)
                    column_name = 'Tag'
                    new_values = ['Update']  # values to update in the column
                    tag = update_excel_column_in_s3(bucket_name, file_key, sheet_name,  column_name, new_values, updated_file_key)                    
                    print(tag)
                else:
                    print("::: Failed to complete script :::")
                    sheet_name = 'input'
                    column_name = 'Processing_Status'
                    new_values = ['The Script is failing. please contact IE Team...!']  # values to update in the column
                    tag = update_excel_column_in_s3(bucket_name, file_key, sheet_name,  column_name, new_values, updated_file_key)
            else:
                    print(f"No request to update data, {tag[0]}")
    except Exception as e:
        print(f"An error occurred while updating Overwritten Impressions: {e}")
        return None

def main(aws_rds_connection,rds_sql_user,rds_sql_pass,rdssql_schema,bucket_name):
    sheet_name = 'input'
    column_name = 'Processing_Status'
    file_key = 'data/inbound/override_backend_utility_input/backend_utility_input.xlsx'
    print(list_sheet_names(bucket_name, file_key))
    updated_file_key = ''  # Optional, save with a new name
    initial_tag = update_excel_column_in_s3(bucket_name, file_key, sheet_name,  column_name, updated_file_key)
    if initial_tag[1] == "Request":
        new_values = ['in-progress']  # values to update in the column
        global connection_properties
        connection_properties = {
            "host": aws_rds_connection,   # replace with your RDS endpoint
            "user": rds_sql_user,
            "password": rds_sql_pass,
            "database": rdssql_schema,
            "port": 3306   
        }
        connection = create_rds_connection(connection_properties)
        global tag
        tag = update_excel_column_in_s3(bucket_name, file_key, sheet_name,  column_name, new_values, updated_file_key)
        print(list_sheet_names(bucket_name, file_key))
        construct_query_pre(connection,bucket_name, file_key, sheet_name, column_name, new_values, updated_file_key)    
        update_levels_event_combinations(connection)
        update_overwrite_for_combinations(connection,bucket_name, file_key, sheet_name, column_name, new_values, updated_file_key)
        generate_download_url(bucket_name, file_key)
    else:
        print("Note: The Status provided in the control file is '{}'. To run the report, please ensure the Status is marked as 'Request'; otherwise, the utility will not execute any of the supported functionalities below.\n".format("Blank" if not initial_tag[1] else initial_tag[1]))
        print("Functionality-1: In case you want to trigger report, please set the 'Tag' column in Sheet 'input' as 'Report'... ")
        print("Functionality-2: In case you want to get all active Levels and Events Combinations, please set the 'Tag' column in Sheet 'input' as 'Active_Levels_and_Event'... ")
        print("Functionality-3: In case you want to update Overwrite Impressions for the combinations Levels and Event, please set the 'Tag' column in Sheet 'input' as 'Update'... ")
        print("Functionality-4: to execute the following utility, user must have to set the Start_Date, End_Date, and {Level1  or Event}. if any of the values are not given the utility may not able to update the overwritten impressions'... \n")
        print("Note: You can download the file from following URL directly aand please ask BAU team to upload the file on below S3 Location.\n")
        print("S3 Location ::: 's3://{}/{}'".format(bucket_name,file_key))
        generate_download_url(bucket_name, file_key)


__name__ = "__main__"
if __name__ == "__main__":
    aws_rds_connection = sys.argv[1]
    rds_sql_user = sys.argv[2]
    rds_sql_pass = sys.argv[3]
    rdssql_schema = 'data_activation'
    bucket_name  = sys.argv[4]#'foxtel-hawkes3-data'  ye-dev-  

    main(aws_rds_connection,rds_sql_user,rds_sql_pass,rdssql_schema,bucket_name)
    
